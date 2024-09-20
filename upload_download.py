import time
import openpyxl

from selenium import webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

driver = None


def driverSetup():
    global driver
    driver = wd.Chrome()
    driver.maximize_window()
    driver.implicitly_wait(5)
    driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
    driver.find_element(By.ID, "downloadButton").click()
    time.sleep(5)

    upload_download()


def upload_download():

    element_extraction()
    execlUtilities()

    file_input = driver.find_element(By.XPATH, "//*[@type='file']")
    file_input.send_keys(r"C:\Users\siddh\OneDrive\Desktop\download.xlsx")

    WebDriverWait(driver, 10).until(
        expected_conditions.visibility_of_element_located((By.XPATH, "//*[text()='Updated Excel Data Successfully.']")))
    msg = driver.find_element(By.XPATH, "//*[text()='Updated Excel Data Successfully.']").text
    print(msg)

    element_extraction()

    time.sleep(3)
    driver.quit()


def execlUtilities():
    book = openpyxl.load_workbook(r"C:\Users\siddh\Downloads\download.xlsx")
    sheet = book.active

    dictFruitname = {}

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if str(sheet.cell(row=i, column=j).value) == "399":
                dictFruitname["row"] = i
                dictFruitname["col"] = j

    sheet.cell(row=dictFruitname["row"], column=dictFruitname["col"]).value = "400"
    book.save(r"C:\Users\siddh\OneDrive\Desktop\download.xlsx")


def element_extraction():
    fruit = "Kivi"
    fruitPath = "//*[text()='" + fruit + "']"

    attribute = driver.find_element(By.XPATH, "//*[text()='Price']").get_attribute("data-column-id")
    pricePath = fruitPath + "/parent::*/parent::*//*[@data-column-id='" + attribute + "']"

    price = driver.find_element(By.XPATH, pricePath).text
    print(price)


driverSetup()
