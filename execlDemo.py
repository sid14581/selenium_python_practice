import openpyxl

book = openpyxl.load_workbook("C:\\Users\\siddh\\OneDrive\\Desktop\\demoData.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)

sheet.cell(row=2, column=2).value = "Sid"

exe_dict = {}
sheet['C3'].value = "harsh@gmail.com"

book.save("C:\\Users\\siddh\\OneDrive\\Desktop\\demoData.xlsx")

# print("##############################################################")

for i in range(2, sheet.max_row + 1):
    for j in range(2, sheet.max_column + 1):

        exe_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print([exe_dict])
